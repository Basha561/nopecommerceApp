import openpyxl
class XLUtils:
    @staticmethod
    def getRowCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_row)

    @staticmethod
    def getColumnCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_column)

    @staticmethod
    def readData(file, sheetName, rowNum, colNum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(rowNum, colNum).value

    @staticmethod
    def writeData(file, sheetName, rowNum, colNum, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(rowNum, colNum).value = data
        workbook.save(file)



